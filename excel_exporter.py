"""
ICS etkinliklerini Excel dosyasına aktaran modül
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont
from openpyxl.utils import get_column_letter
from typing import List, Dict
import os
from datetime import datetime


class ExcelExporter:
    """ICS etkinliklerini Excel dosyasına aktaran sınıf"""
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def export_to_excel(self, events: List[Dict], output_path: str, sheet_name: str = "Events"):
        """
        Etkinlikleri Excel dosyasına aktarır
        
        Args:
            events: Etkinlik listesi
            output_path: Çıktı Excel dosyası yolu
            sheet_name: Sheet adı (varsayılan: "Events")
        """
        if not events:
            raise ValueError("Yüklenecek etkinlik bulunamadı")
        
        # Workbook oluştur
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = sheet_name
        
        # Başlık satırı
        headers = ['Başlık', 'Başlangıç', 'Bitiş', 'Açıklama', 'Konum', 'Organizer', 'URL', 'UID']
        
        # Başlıkları yaz ve formatla
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True, size=11)
        
        for col_num, header in enumerate(headers, 1):
            cell = self.worksheet.cell(row=1, column=col_num)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Veri satırlarını yaz (formatlanmış metin desteği ile)
        for row_num, event in enumerate(events, 2):
            # Başlık (formatlanmış)
            summary_cell = self.worksheet.cell(row=row_num, column=1)
            self._set_formatted_text(summary_cell, event.get('summary_formatted', []), event.get('summary', ''))
            
            # Diğer alanlar (düz metin)
            self.worksheet.cell(row=row_num, column=2, value=event.get('start', ''))
            self.worksheet.cell(row=row_num, column=3, value=event.get('end', ''))
            
            # Açıklama (formatlanmış)
            desc_cell = self.worksheet.cell(row=row_num, column=4)
            self._set_formatted_text(desc_cell, event.get('description_formatted', []), event.get('description', ''))
            
            self.worksheet.cell(row=row_num, column=5, value=event.get('location', ''))
            self.worksheet.cell(row=row_num, column=6, value=event.get('organizer', ''))
            self.worksheet.cell(row=row_num, column=7, value=event.get('url', ''))
            self.worksheet.cell(row=row_num, column=8, value=event.get('uid', ''))
        
        # Sütun genişliklerini ayarla
        column_widths = {
            'A': 30,  # Başlık
            'B': 20,  # Başlangıç
            'C': 20,  # Bitiş
            'D': 40,  # Açıklama
            'E': 25,  # Konum
            'F': 30,  # Organizer
            'G': 40,  # URL
            'H': 40   # UID
        }
        
        for col_letter, width in column_widths.items():
            self.worksheet.column_dimensions[col_letter].width = width
        
        # Satır yüksekliklerini ayarla
        self.worksheet.row_dimensions[1].height = 25  # Başlık satırı
        
        # Hücreleri wrap text yap (uzun metinler için)
        for row in self.worksheet.iter_rows(min_row=2, max_row=len(events) + 1):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        
        # Dosyayı kaydet
        self.workbook.save(output_path)
    
    def _set_formatted_text(self, cell, formatted_parts: List[Dict], fallback_text: str):
        """
        Hücreye formatlanmış metin yazar (kalın, italik vb.)
        
        Args:
            cell: Excel hücresi
            formatted_parts: Formatlanmış metin parçaları listesi
            fallback_text: Formatlanmış parça yoksa kullanılacak düz metin
        """
        if not formatted_parts or not any(part.get('text', '').strip() for part in formatted_parts):
            # Formatlanmış parça yoksa düz metin kullan
            cell.value = fallback_text
            return
        
        # RichText oluştur
        rich_text_parts = []
        for part in formatted_parts:
            text = part.get('text', '')
            if not text:
                continue
            
            # InlineFont ayarları (TextBlock için InlineFont gerekli)
            font_kwargs = {}
            if part.get('bold', False):
                font_kwargs['b'] = True  # bold için 'b' kullanılır
            if part.get('italic', False):
                font_kwargs['i'] = True  # italic için 'i' kullanılır
            
            if font_kwargs:
                inline_font = InlineFont(**font_kwargs)
                rich_text_parts.append(TextBlock(inline_font, text))
            else:
                inline_font = InlineFont()
                rich_text_parts.append(TextBlock(inline_font, text))
        
        if rich_text_parts:
            try:
                cell.value = CellRichText(rich_text_parts)
            except:
                # RichText desteklenmiyorsa düz metin kullan
                cell.value = fallback_text
        else:
            cell.value = fallback_text
    
    def export_to_csv(self, events: List[Dict], output_path: str):
        """
        Etkinlikleri CSV dosyasına aktarır (alternatif format)
        
        Args:
            events: Etkinlik listesi
            output_path: Çıktı CSV dosyası yolu
        """
        import csv
        
        if not events:
            raise ValueError("Yüklenecek etkinlik bulunamadı")
        
        headers = ['Başlık', 'Başlangıç', 'Bitiş', 'Açıklama', 'Konum', 'Organizer', 'URL', 'UID']
        
        with open(output_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
            writer = csv.writer(csvfile, delimiter=';')
            writer.writerow(headers)
            
            for event in events:
                row = [
                    event.get('summary', ''),
                    event.get('start', ''),
                    event.get('end', ''),
                    event.get('description', ''),
                    event.get('location', ''),
                    event.get('organizer', ''),
                    event.get('url', ''),
                    event.get('uid', '')
                ]
                writer.writerow(row)
